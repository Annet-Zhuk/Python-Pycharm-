from bs4 import BeautifulSoup
from urllib.request import urlopen
import urllib.request
import openpyxl
from openpyxl import load_workbook
import json
wb = load_workbook('task_9_4_1_data-25290-2019-09-30.xlsx')
sheet = wb['Sheet0']
d = {}
sAdm = []
sDist = []
jon = {}
for i in range (2, sheet.max_row + 1):
   District = sheet.cell(row=i, column=6).value
   AdmArea = sheet.cell(row=i, column=5).value
   if AdmArea not in sAdm:
      sAdm.append(AdmArea)
for Adm in sAdm:
   k = {}
   sDis = []
   for i in range(2, sheet.max_row + 1):
      AdmArea1 = sheet.cell(row=i, column=5).value
      District2 = sheet.cell(row=i, column=6).value
      if Adm == AdmArea1:
         sDis.append(District2)
   for Dis in sDis:
      sAdd = []
      for i in range(2, sheet.max_row + 1):
         District1 = sheet.cell(row=i, column=6).value
         Address = sheet.cell(row=i, column=7).value
         if Dis == District1:
            sAdd.append(Address)
      d[Dis] = sAdd
   jon[Adm] = d
print(json.dumps(jon, ensure_ascii=False))
