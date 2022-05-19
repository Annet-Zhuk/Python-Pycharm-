import openpyxl
from openpyxl import load_workbook
wb = load_workbook('trekking2_6_6_2.xlsx')
sheet = wb['Справочник']
d = {}
for i in range (2, sheet.max_row + 1):
   name = sheet.cell(row=i, column=1).value
   kkal = sheet.cell(row=i, column=2).value
   B = sheet.cell(row=i, column=3).value
   G = sheet.cell(row=i, column=4).value
   U = sheet.cell(row=i, column=5).value
   d[name] = kkal, B, G, U

sheet2 = wb['Раскладка']
r = {}
kal_sum_day_ = 0
proteins_sum_day_ = 0
fats_sum_day_ = 0
carbohydrates_sum_day_ = 0

for i in range (2, sheet2.max_row + 1):
   produkt = sheet2.cell(row=i, column=1).value
   weight = sheet2.cell(row=i, column=2).value
   kal_sum = 0
   proteins_sum = 0
   fats_sum = 0
   carbohydrates_sum = 0
   kal_sum = float(d[produkt][0]) * int(weight) / 100 + kal_sum  # каллорийность
   proteins_sum = float(d[produkt][1]) * int(weight) / 100 + proteins_sum  # белки
   fats_sum = float(d[produkt][2]) * int(weight) / 100 + fats_sum  # жиры
   if (d[produkt][3]) != None:
      carbohydrates_sum = float((d[produkt][3]) * int(weight) / 100 + carbohydrates_sum)

   kal_sum_day_ = kal_sum_day_ + kal_sum
   proteins_sum_day_ = proteins_sum_day_ + proteins_sum
   fats_sum_day_ = fats_sum_day_ + fats_sum
   carbohydrates_sum_day_ = carbohydrates_sum_day_ + carbohydrates_sum

print(int(kal_sum_day_), end=" ")
print(int(proteins_sum_day_), end=" ")
print(int(fats_sum_day_), end=" ")
print(int(carbohydrates_sum_day_))

